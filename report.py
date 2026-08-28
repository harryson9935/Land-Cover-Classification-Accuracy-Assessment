"""
report.py
----------
Generates all deliverable artifacts:
  - comparison_results.xlsx : full grid-search comparison + best-config
                               confusion matrix & per-class accuracy sheets
  - accuracy_report.html    : self-contained HTML accuracy assessment report
  - PNG images               : RGB composite, ground-truth map, classified
                               map, confusion matrix heatmap, feature
                               importance chart
"""

import os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from data_generator import CLASS_NAMES, scene_to_rgb
from classifier import predict_full_scene, feature_importances

CLASS_COLORS = [
    "#419BDF", "#397D49", "#88B053", "#7A87C6", "#E49635",
    "#DFC35A", "#C4281B", "#A59B8F", "#B39FE1",
]


def save_scene_images(image, labels, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    cmap = ListedColormap(CLASS_COLORS)

    rgb = scene_to_rgb(image)
    plt.figure(figsize=(6, 6))
    plt.imshow(rgb)
    plt.title("Sentinel-2-style RGB composite (synthetic scene)")
    plt.axis("off")
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, "01_rgb_composite.png"), dpi=150)
    plt.close()

    plt.figure(figsize=(7, 6))
    im = plt.imshow(labels, cmap=cmap, vmin=0, vmax=len(CLASS_NAMES) - 1)
    plt.title("Ground truth (Dynamic World-style labels)")
    plt.axis("off")
    cbar = plt.colorbar(im, ticks=range(len(CLASS_NAMES)), fraction=0.046, pad=0.04)
    cbar.ax.set_yticklabels(CLASS_NAMES)
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, "02_ground_truth_map.png"), dpi=150)
    plt.close()
    return cmap


def save_classified_map(labels_pred, cmap, out_dir, title, filename):
    plt.figure(figsize=(7, 6))
    im = plt.imshow(labels_pred, cmap=cmap, vmin=0, vmax=len(CLASS_NAMES) - 1)
    plt.title(title)
    plt.axis("off")
    cbar = plt.colorbar(im, ticks=range(len(CLASS_NAMES)), fraction=0.046, pad=0.04)
    cbar.ax.set_yticklabels(CLASS_NAMES)
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, filename), dpi=150)
    plt.close()


def save_confusion_matrix_plot(cm, class_names, out_path, title="Confusion Matrix (best configuration)"):
    cm_norm = cm.astype(float) / (cm.sum(axis=1, keepdims=True) + 1e-9)
    fig, ax = plt.subplots(figsize=(8, 7))
    im = ax.imshow(cm_norm, cmap="Blues", vmin=0, vmax=1)
    ax.set_xticks(range(len(class_names)))
    ax.set_yticks(range(len(class_names)))
    ax.set_xticklabels(class_names, rotation=45, ha="right")
    ax.set_yticklabels(class_names)
    ax.set_xlabel("Predicted class")
    ax.set_ylabel("Reference (true) class")
    ax.set_title(title)
    for i in range(len(class_names)):
        for j in range(len(class_names)):
            val = cm[i, j]
            color = "white" if cm_norm[i, j] > 0.5 else "black"
            ax.text(j, i, str(val), ha="center", va="center", color=color, fontsize=8)
    fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04, label="Row-normalized proportion")
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()


def save_feature_importance_plot(clf, feature_names, out_path, top_n=15):
    top = feature_importances(clf, feature_names, top_n=top_n)
    names = [t[0] for t in top][::-1]
    vals = [t[1] for t in top][::-1]
    plt.figure(figsize=(7, 6))
    plt.barh(names, vals, color="#397D49")
    plt.xlabel("RF feature importance (Gini)")
    plt.title(f"Top {top_n} features — best configuration")
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()


def save_comparison_chart(df, out_path):
    fig, ax1 = plt.subplots(figsize=(9, 5))
    labels_x = [f"w{r.window_size}\n{r.sampling_strategy}" for r in df.itertuples()]
    x = np.arange(len(df))
    ax1.bar(x - 0.18, df["overall_accuracy_%"], width=0.36, label="Overall Accuracy (%)", color="#397D49")
    ax1.set_ylabel("Overall Accuracy (%)")
    ax1.set_ylim(0, 100)
    ax2 = ax1.twinx()
    ax2.bar(x + 0.18, df["kappa"] * 100, width=0.36, label="Kappa (x100)", color="#E49635")
    ax2.set_ylabel("Kappa (x100)")
    ax2.set_ylim(0, 100)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels_x, fontsize=8)
    ax1.set_title("Accuracy across window sizes & sampling strategies")
    lines1, labs1 = ax1.get_legend_handles_labels()
    lines2, labs2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labs1 + labs2, loc="upper center",
               bbox_to_anchor=(0.5, -0.18), ncol=2, frameon=False)
    plt.subplots_adjust(bottom=0.28)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150)
    plt.close()


def export_excel(df, best_result, class_names, out_path):
    per_class = best_result["_full_result"]["per_class"]
    cm = best_result["_full_result"]["confusion_matrix"]

    per_class_df = pd.DataFrame([
        {
            "class": c,
            "support_(test_px)": per_class[c]["support"],
            "producers_accuracy_%": round(per_class[c]["producers_accuracy"] * 100, 2),
            "users_accuracy_%": round(per_class[c]["users_accuracy"] * 100, 2),
            "f1_score": round(per_class[c]["f1_score"], 3),
        }
        for c in class_names
    ])

    cm_df = pd.DataFrame(cm, index=[f"true_{c}" for c in class_names],
                          columns=[f"pred_{c}" for c in class_names])

    summary_df = pd.DataFrame([{
        "metric": "Best configuration",
        "value": f"window_size={best_result['window_size']}, "
                 f"sampling={best_result['sampling_strategy']}"
    }, {
        "metric": "Overall Accuracy (%)",
        "value": round(best_result["overall_accuracy"] * 100, 2)
    }, {
        "metric": "Kappa coefficient",
        "value": round(best_result["kappa"], 3)
    }, {
        "metric": "Macro-averaged F1",
        "value": round(best_result["macro_f1"], 3)
    }, {
        "metric": "Number of training pixels",
        "value": best_result["n_train"]
    }, {
        "metric": "Number of test pixels",
        "value": best_result["n_test"]
    }, {
        "metric": "Number of features",
        "value": best_result["n_features"]
    }])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="Experiment_Comparison", index=False)
        per_class_df.to_excel(writer, sheet_name="Per_Class_Accuracy", index=False)
        cm_df.to_excel(writer, sheet_name="Confusion_Matrix")

    _style_workbook(out_path)


def _style_workbook(path):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="397D49", end_color="397D49", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    body_font = Font(name="Arial")

    for ws in wb.worksheets:
        max_col = ws.max_column
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = max(
                [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)]
                + [10]
            )
            ws.column_dimensions[letter].width = min(max_len + 3, 40)
        ws.freeze_panes = "A2"
    wb.save(path)


def export_html(df, best_result, class_names, image_dir, out_path):
    per_class = best_result["_full_result"]["per_class"]
    rows_html = "\n".join(
        f"<tr><td>{c}</td><td>{per_class[c]['support']}</td>"
        f"<td>{per_class[c]['producers_accuracy']*100:.2f}%</td>"
        f"<td>{per_class[c]['users_accuracy']*100:.2f}%</td>"
        f"<td>{per_class[c]['f1_score']:.3f}</td></tr>"
        for c in class_names
    )
    comparison_rows = "\n".join(
        f"<tr><td>{r.window_size}</td><td>{r.sampling_strategy}</td>"
        f"<td>{r.n_features}</td><td>{r.overall_accuracy_percent}</td>"
        f"<td>{r.kappa}</td><td>{r.macro_f1}</td><td>{r.train_time_sec}</td></tr>"
        for r in df.rename(columns={"overall_accuracy_%": "overall_accuracy_percent"}).itertuples()
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Land Cover Classification — Accuracy Assessment Report</title>
<style>
  body {{ font-family: Arial, Helvetica, sans-serif; margin: 40px; color: #222; background:#fafafa;}}
  h1 {{ color: #2c5530; border-bottom: 3px solid #397D49; padding-bottom: 8px;}}
  h2 {{ color: #2c5530; margin-top: 40px; }}
  table {{ border-collapse: collapse; width: 100%; margin: 16px 0; background: white;}}
  th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: center; }}
  th {{ background-color: #397D49; color: white; }}
  tr:nth-child(even) {{ background-color: #f2f6f2; }}
  .metric-cards {{ display: flex; gap: 16px; flex-wrap: wrap; margin: 20px 0; }}
  .card {{ background: white; border-left: 5px solid #397D49; padding: 16px 24px; border-radius: 6px;
           box-shadow: 0 1px 4px rgba(0,0,0,0.08); min-width: 180px;}}
  .card .value {{ font-size: 28px; font-weight: bold; color: #2c5530; }}
  .card .label {{ font-size: 13px; color: #666; text-transform: uppercase; letter-spacing: 0.03em;}}
  img {{ max-width: 100%; border: 1px solid #ddd; border-radius: 6px; margin: 10px 0; }}
  .imgrow {{ display:flex; gap:16px; flex-wrap:wrap; }}
  .imgrow figure {{ flex: 1; min-width: 260px; margin:0;}}
  figcaption {{ font-size: 13px; color:#555; text-align:center; margin-top:4px;}}
  .note {{ background:#fff8e1; border-left:5px solid #E49635; padding:12px 18px; border-radius:6px; font-size:14px;}}
  footer {{ margin-top:50px; font-size:12px; color:#888; }}
</style>
</head>
<body>
<h1>Land Cover Classification — Accuracy Assessment Report</h1>
<p>Sentinel-2 multispectral imagery &middot; Dynamic World 9-class schema &middot; Random Forest, window-based spectral-spatial features</p>

<div class="note"><strong>Data note:</strong> this run uses a synthetically generated Sentinel-2 / Dynamic-World-style
scene (see <code>src/data_generator.py</code>) because this environment has no internet access to pull real
imagery from Google Earth Engine. Swap in real Sentinel-2 + Dynamic World rasters and the rest of the pipeline
runs unchanged.</div>

<h2>Best Configuration</h2>
<div class="metric-cards">
  <div class="card"><div class="value">{best_result['overall_accuracy']*100:.2f}%</div><div class="label">Overall Accuracy</div></div>
  <div class="card"><div class="value">{best_result['kappa']:.3f}</div><div class="label">Kappa Coefficient</div></div>
  <div class="card"><div class="value">{best_result['macro_f1']:.3f}</div><div class="label">Macro F1</div></div>
  <div class="card"><div class="value">window={best_result['window_size']}</div><div class="label">Best Window Size</div></div>
  <div class="card"><div class="value">{best_result['sampling_strategy']}</div><div class="label">Sampling Strategy</div></div>
</div>

<h2>Per-Class Accuracy</h2>
<table>
<tr><th>Class</th><th>Support (test px)</th><th>Producer's Accuracy</th><th>User's Accuracy</th><th>F1 Score</th></tr>
{rows_html}
</table>

<h2>Window Size &times; Sampling Strategy Comparison</h2>
<table>
<tr><th>Window Size</th><th>Sampling</th><th>#Features</th><th>OA (%)</th><th>Kappa</th><th>Macro F1</th><th>Train Time (s)</th></tr>
{comparison_rows}
</table>

<h2>Visual Results</h2>
<div class="imgrow">
  <figure><img src="images/01_rgb_composite.png"><figcaption>Sentinel-2-style RGB composite</figcaption></figure>
  <figure><img src="images/02_ground_truth_map.png"><figcaption>Ground truth (Dynamic World labels)</figcaption></figure>
  <figure><img src="images/03_classified_map.png"><figcaption>RF classified map (best configuration)</figcaption></figure>
</div>
<div class="imgrow">
  <figure><img src="images/04_confusion_matrix.png"><figcaption>Confusion matrix (row-normalized)</figcaption></figure>
  <figure><img src="images/05_feature_importance.png"><figcaption>Top feature importances</figcaption></figure>
  <figure><img src="images/06_comparison_chart.png"><figcaption>Accuracy vs. window size / sampling strategy</figcaption></figure>
</div>

<footer>Generated automatically by report.py — Land Cover Classification &amp; Accuracy Assessment pipeline.</footer>
</body>
</html>"""
    with open(out_path, "w") as f:
        f.write(html)
